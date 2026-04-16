from fastapi import FastAPI, Depends, HTTPException, status, WebSocket, WebSocketDisconnect, Request, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, text
from typing import List, Dict, Optional
import asyncio
import datetime
import io
import logging
import os
from contextlib import asynccontextmanager
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.units import cm
from apscheduler.schedulers.background import BackgroundScheduler
from notifications import send_supplier_email, trigger_stock_webhook, send_admin_report

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger("akilli_envanter_api")

# --- UYGULAMA SABİTLERİ ---
SKT_ALERT_DAYS = 7                    # SKT yaklaşma uyarı eşiği (gün)
AUDIT_LOG_ARCHIVE_DAYS = 7            # Denetim kaydı arşiv süresi (gün)
REPORTING_WINDOW_DAYS = 30            # Standart raporlama penceresi (gün)
EWMA_ALPHA = 0.3                      # EWMA talep tahmini alfa katsayısı

import models
import schemas
from database import engine, get_db
from auth import (
    verify_password, get_password_hash, create_access_token,
    create_refresh_token, decode_refresh_token,
    role_required, get_current_user,
)
from schemas import StockTransaction, ProductCreate, ProductUpdate, TalepYaniti, ProductLifecycleResponse, LifecycleSummary, LifecycleEvent



class RegisterUser(BaseModel):
    username: str
    password: str
    role: str = "Barista"

class SupplyApprovalRequest(BaseModel):
    product_id: int
    quantity: float
    supplier_name: Optional[str] = None

# Not: Veritabanı tabloları artık Alembic (Migrations) ile yönetiliyor.
# Manuel oluşturma (create_all) devri enterprise mimaride kapandı.

# ==========================================
# --- RATE LIMITER (BRUTE-FORCE KORUMASI) ---
# ==========================================
limiter = Limiter(key_func=get_remote_address)

# ==========================================
# --- 🕰️ PLANLANMIŞ GÖREVLER (CRON) ---
# ==========================================
scheduler = BackgroundScheduler()

def daily_system_check():
    """Her gece 00:00'da çalışan kritik kontroller."""
    db = next(get_db())
    try:
        # 1. SKT Yaklaşanlar (7 gün)
        bugun = datetime.date.today()
        kritik_tarih = bugun + datetime.timedelta(days=SKT_ALERT_DAYS)
        skt_yaklasanlar = db.query(models.Product).filter(
            models.Product.expiration_date <= kritik_tarih,
            models.Product.current_stock > 0
        ).all()
        
        # 2. Kritik Stok Özeti
        kritik_stoklar = db.query(models.Product).filter(
            models.Product.current_stock <= models.Product.reorder_point
        ).all()

        if skt_yaklasanlar or kritik_stoklar:
            report = f"📅 Günlük Sistem Raporu - {bugun}\n\n"
            if skt_yaklasanlar:
                report += "⚠️ SKT Yaklaşan Ürünler:\n"
                for u in skt_yaklasanlar:
                    report += f"- {u.name_tr} (SKT: {u.expiration_date}, Stok: {u.current_stock})\n"
            
            if kritik_stoklar:
                report += "\n📉 Kritik Stok Seviyesindeki Ürünler:\n"
                for u in kritik_stoklar:
                    report += f"- {u.name_tr} (Mevcut: {u.current_stock}, Eşik: {u.reorder_point})\n"
            
            asyncio.run(send_admin_report(f"🏛️ Envanter Raporu - {bugun}", report))
            logger.info("Günlük sistem raporu hazırlandı ve kuyruğa alındı.")

    except Exception as e:
        logger.error(f"Planlanmış görev hatası: {str(e)}")
    finally:
        db.close()

def archive_old_audit_logs():
    """Her gece 01:00'de 7 günden eski denetim kayıtlarını arşivler (siler değil, gizler)."""
    db = next(get_db())
    try:
        cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=AUDIT_LOG_ARCHIVE_DAYS)
        eski_kayitlar = db.query(models.AuditLog).filter(
            models.AuditLog.timestamp < cutoff,
            models.AuditLog.is_archived == 0
        ).all()
        for log in eski_kayitlar:
            log.is_archived = 1
            log.archived_at = datetime.datetime.utcnow()
        db.commit()
        logger.info(f"{len(eski_kayitlar)} denetim kaydı arşivlendi.")
    except Exception as e:
        logger.error(f"Arshiv görevi hatası: {str(e)}")
    finally:
        db.close()

def _ensure_is_approved_column():
    """Lightweight migration: app_users.is_approved sütununu ekler, mevcut kullanıcıları onaylı sayar."""
    from sqlalchemy import text
    from database import engine
    try:
        with engine.begin() as conn:
            cols = conn.execute(text("SHOW COLUMNS FROM app_users LIKE 'is_approved'")).fetchall()
            if not cols:
                conn.execute(text("ALTER TABLE app_users ADD COLUMN is_approved INT DEFAULT 1"))
                conn.execute(text("UPDATE app_users SET is_approved = 1"))
                logger.info("app_users.is_approved sütunu eklendi; mevcut kullanıcılar onaylı olarak işaretlendi.")
    except Exception as e:
        logger.error(f"is_approved migration hatası: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    _ensure_is_approved_column()
    logger.info("İş Zekası Envanter API başarıyla başlatıldı.")
    yield
    scheduler.shutdown()
    logger.info("API kapatılıyor.")

# Scheduler görevlerini başlat
scheduler.add_job(daily_system_check, 'cron', hour=0, minute=0, id='daily_check')
scheduler.add_job(archive_old_audit_logs, 'cron', hour=1, minute=0, id='audit_archive')
scheduler.start()

app = FastAPI(
    title="Akıllı Kafe Envanter Sistemi",
    version="2.1",
    lifespan=lifespan
)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Rate Limit handler
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

allowed_origins = os.getenv("ALLOWED_ORIGINS", "*").split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins, 
    allow_credentials=True,
    allow_methods=["*"], 
    allow_headers=["*"],
)

# --- YARDIMCI: İSTEK'TEN IP ADRESİ ÇEK ---
def get_client_ip(request: Request) -> str:
    """Proxy arkasında bile doğru IP adresini al (X-Forwarded-For)."""
    forwarded = request.headers.get("X-Forwarded-For")
    return forwarded.split(",")[0].strip() if forwarded else (request.client.host if request.client else "unknown")

# --- YARDIMCI: DENETİM KAYDI OLUŞTUR ---
def _log_audit(
    db: Session,
    actor: str,
    role: str,
    action: str,
    resource: str = None,
    detail: str = None,
    ip_address: str = None,
) -> None:
    """Denetim kaydı yazar. Hata olursa sessizce geçmek yerine loglar."""
    try:
        log = models.AuditLog(
            actor=actor, role=role, action=action,
            resource=resource, detail=detail, ip_address=ip_address,
        )
        db.add(log)
        db.commit()
    except Exception as e:
        logger.error(f"Audit log yazılamadı [{action}]: {e}")

# --- WEBSOCKET CANLI YAYIN (BROADCAST) YÖNETİCİSİ ---
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            await connection.send_text(message)

manager = ConnectionManager()

@app.websocket("/ws/notifications")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Bekleme döngüsü (Bağlantıyı canlı tutar)
            data = await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)




# ==========================================
# --- 0. SİSTEM GİRİŞİ (LOGIN & JWT TOKEN) ---
# ==========================================
@app.post("/token")
@limiter.limit("10/minute")  # Brute-force koruması: dakikada max 10 deneme
def login_for_access_token(request: Request, form_data: dict, db: Session = Depends(get_db)):
    username = form_data.get("username")
    password = form_data.get("password")
    ip = get_client_ip(request)
    
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user or not verify_password(password, user.hashed_password):
        _log_audit(db, actor=username or "unknown", role="-",
                   action="LOGIN_FAILED", detail="Yanlış şifre denendi.", ip_address=ip)
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Kullanıcı adı veya şifre hatalı. Lütfen tekrar deneyiniz.",
            headers={"WWW-Authenticate": "Bearer"},
        )

    if user.role != "Admin" and not getattr(user, "is_approved", 0):
        _log_audit(db, actor=user.username, role=user.role,
                   action="LOGIN_BLOCKED", detail="Hesap admin onayı bekliyor.", ip_address=ip)
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Hesabınız henüz admin tarafından onaylanmadı. Lütfen onay bekleyiniz.",
        )
    
    access_token  = create_access_token(data={"sub": user.username})
    refresh_token = create_refresh_token(data={"sub": user.username})

    _log_audit(db, actor=user.username, role=user.role,
               action="LOGIN", detail="JWT access + refresh token issued.", ip_address=ip)

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "role": user.role,
        "expires_in_minutes": 120
    }


@app.post("/auth/refresh")
@limiter.limit("20/minute")
def refresh_access_token(request: Request, body: dict, db: Session = Depends(get_db)):
    """Refresh token ile yeni bir access token üretir. Frontend login olmadan devam edebilir."""
    refresh_token = body.get("refresh_token")
    if not refresh_token:
        raise HTTPException(status_code=400, detail="refresh_token gerekli.")
    
    username = decode_refresh_token(refresh_token)  # geçersizse 401 fırlatır
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user:
        raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı.")
    
    new_access = create_access_token(data={"sub": user.username})
    logger.info(f"Token yenilendi: {username} | IP: {get_client_ip(request)}")
    return {"access_token": new_access, "token_type": "bearer", "expires_in_minutes": 120}

# --- 0.5 SİSTEME KAYIT (YENİ PERSONEL) ---
@app.post("/register")
def register_user(user_data: RegisterUser, db: Session = Depends(get_db)):
    mevcut = db.query(models.User).filter(models.User.username == user_data.username).first()
    if mevcut:
        raise HTTPException(status_code=409, detail="Bu kullanıcı adı sistemde zaten kayıtlı! Lütfen farklı bir isim seçin.")

    role = user_data.role if user_data.role in ["Admin", "Depo Müdürü", "Barista"] else "Barista"
    # Admin dışındaki tüm yeni kayıtlar onay bekler.
    is_approved = 1 if role == "Admin" else 0
    try:
        yeni_kullanici = models.User(
            username=user_data.username,
            hashed_password=get_password_hash(user_data.password),
            role=role,
            is_approved=is_approved,
        )
        db.add(yeni_kullanici)
        db.commit()
        if is_approved:
            return {"mesaj": f"Sisteme '{user_data.username}' rolüyle '{role}' olarak kaydedildiniz. Giriş yapabilirsiniz."}
        return {"mesaj": f"Kayıt alındı. Hesabınız admin onayını bekliyor — onaylandığında giriş yapabilirsiniz."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Bir iç sistem hatası oluştu: {str(e)}")


# --- 0.6 ADMIN: KULLANICI ONAY YÖNETİMİ ---
@app.get("/pending-users", dependencies=[Depends(role_required(["Admin"]))])
def pending_users_list(db: Session = Depends(get_db)):
    users = db.query(models.User).filter(models.User.is_approved == 0).all()
    return {"users": [{"id": u.id, "username": u.username, "role": u.role} for u in users]}


@app.post("/approve-user/{user_id}", dependencies=[Depends(role_required(["Admin"]))])
def approve_user(user_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı.")
    user.is_approved = 1
    db.commit()
    _log_audit(db, actor=current_user.username, role=current_user.role,
               action="USER_APPROVED", resource=user.username, detail=f"Rol: {user.role}")
    return {"mesaj": f"{user.username} onaylandı."}


@app.delete("/reject-user/{user_id}", dependencies=[Depends(role_required(["Admin"]))])
def reject_user(user_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == user_id, models.User.is_approved == 0).first()
    if not user:
        raise HTTPException(status_code=404, detail="Bekleyen kullanıcı bulunamadı.")
    uname = user.username
    db.delete(user)
    db.commit()
    _log_audit(db, actor=current_user.username, role=current_user.role,
               action="USER_REJECTED", resource=uname)
    return {"mesaj": f"{uname} reddedildi ve silindi."}


# ==========================================
# --- 1. SİSTEMİN KALBİ (UÇ NOKTALAR) ---
# ==========================================

@app.get("/urunler", dependencies=[Depends(role_required(["Admin", "Depo Müdürü", "Barista"]))])
def urunleri_getir(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    urunler = db.query(models.Product).offset(skip).limit(limit).all()
    sonuc = []
    # ORM Sayesinde Kategoriler InnerJoin olmaksızın .category üzerinden çağrılabiliyor. Muazzam!
    for u in urunler:
        sonuc.append({
            "product_id": u.product_id,
            "sku": u.sku,
            "name_tr": u.name_tr,
            "name_en": u.name_en,
            "description_tr": u.description_tr,
            "description_en": u.description_en,
            "current_stock": u.current_stock,
            "unit_price": u.unit_price,
            "category_id": u.category_id,
            "supplier_id": u.supplier_id,
            "category_name_tr": u.category.name_tr if u.category else "Tanımsız",
            "category_name_en": u.category.name_en if u.category else "Undefined",
            "supplier_name": u.supplier.name if u.supplier else "Tanımsız",
            "unit_cost": u.unit_cost,
            "reorder_point": u.reorder_point,
            "abc_class": u.abc_class,
            "expiration_date": u.expiration_date.isoformat() if u.expiration_date else None,
            "warehouse_location": u.warehouse_location
        })
    return {"data": sonuc}

@app.get("/")
@app.get("/index.html")
def root():
    return FileResponse("index.html")

@app.get("/sw.js")
def service_worker():
    return FileResponse("sw.js", media_type="application/javascript")




@app.post("/urun-ekle", status_code=201)
def urun_ekle(urun: ProductCreate, current_user: models.User = Depends(role_required(["Admin", "Depo Müdürü"])), db: Session = Depends(get_db)):
    # Strict uniqueness guard — never upsert
    mevcut = db.query(models.Product).filter(models.Product.sku == urun.sku).first()
    if mevcut:
        raise HTTPException(status_code=409, detail=f"Bu SKU zaten mevcut: {urun.sku} / SKU already exists: {urun.sku}")

    try:
        # 1. Product INSERT
        urun_dict = urun.model_dump()
        yeni_urun = models.Product(**urun_dict)
        db.add(yeni_urun)
        db.flush()  # get product_id without committing

        # 2. Initial stock-in movement (uses existing InventoryTransaction table)
        if urun.current_stock > 0:
            hareket = models.InventoryTransaction(
                product_id=yeni_urun.product_id,
                quantity=urun.current_stock,
                transaction_type="IN",
                notes="Yeni ürün envantere eklendi / New product added to inventory",
                processed_by=current_user.username,
                status="ONAYLANDI",
                source="initial_entry",
            )
            db.add(hareket)

        # 3. Audit log (still inside same transaction)
        _log_audit(db, actor=current_user.username, role=current_user.role,
                   action="PRODUCT_CREATE", resource=urun.sku,
                   detail=f"SKU: {urun.sku} | Ad: {urun.name_tr} | Stok: {urun.current_stock}")

        # Single commit — all three writes succeed or all roll back
        db.commit()
        return {"mesaj": f"{urun.name_tr} veritabanına eklendi.", "product_id": yeni_urun.product_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ekleme hatası: {str(e)}")


@app.put("/urun/{product_id}")
def urun_guncelle(product_id: int, urun: ProductUpdate, current_user: models.User = Depends(role_required(["Admin", "Depo Müdürü"])), db: Session = Depends(get_db)):
    mevcut = db.query(models.Product).filter(models.Product.product_id == product_id).first()
    if not mevcut:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı.")
    degisenler = urun.dict(exclude_unset=True)
    for alan, deger in degisenler.items():
        setattr(mevcut, alan, deger)
    db.commit()
    _log_audit(db, actor=current_user.username, role=current_user.role,
               action="PRODUCT_UPDATE", resource=mevcut.name_tr,
               detail=", ".join(f"{k}={v}" for k, v in degisenler.items()))
    return {"mesaj": f"{mevcut.name_tr} güncellendi."}


@app.delete("/urun/{product_id}")
def urun_sil(product_id: int, current_user: models.User = Depends(role_required(["Admin"])), db: Session = Depends(get_db)):
    mevcut = db.query(models.Product).filter(models.Product.product_id == product_id).first()
    if not mevcut:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı.")
    ad = mevcut.name_tr
    db.query(models.InventoryTransaction).filter(models.InventoryTransaction.product_id == product_id).delete(synchronize_session=False)
    db.query(models.SupplyOrderApproval).filter(models.SupplyOrderApproval.product_id == product_id).delete(synchronize_session=False)
    db.delete(mevcut)
    db.commit()
    _log_audit(db, actor=current_user.username, role=current_user.role,
               action="PRODUCT_DELETE", resource=ad, detail=f"ID: {product_id}")
    return {"mesaj": f"{ad} silindi."}


@app.get("/stok-log", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def stok_hareket_log(limit: int = 50, db: Session = Depends(get_db)):
    """Son stok hareketlerini (IN/OUT) aktör bilgisi ile döner."""
    hareketler = (
        db.query(models.InventoryTransaction)
        .order_by(models.InventoryTransaction.transaction_date.desc())
        .limit(limit)
        .all()
    )
    sonuc = []
    for h in hareketler:
        urun_adi = h.product.name_tr if h.product else f"ID:{h.product_id}"
        sku = h.product.sku if h.product else "-"
        sonuc.append({
            "transaction_id": h.transaction_id,
            "sku": sku,
            "urun_adi": urun_adi,
            "quantity": h.quantity,
            "transaction_type": h.transaction_type,
            "processed_by": h.processed_by,
            "status": h.status,
            "notes": h.notes,
            "transaction_date": h.transaction_date.isoformat() if h.transaction_date else None,
        })
    return {"log": sonuc}


@app.get("/api/products/{sku}/lifecycle", response_model=ProductLifecycleResponse)
def product_lifecycle(sku: str, current_user: models.User = Depends(role_required(["Admin", "Depo Müdürü"])), db: Session = Depends(get_db)):
    """Aggregate lifecycle timeline for a single product by SKU."""
    urun = db.query(models.Product).filter(models.Product.sku == sku).first()
    if not urun:
        raise HTTPException(status_code=404, detail=f"SKU '{sku}' bulunamadı.")

    hareketler = (
        db.query(models.InventoryTransaction)
        .filter(models.InventoryTransaction.product_id == urun.product_id)
        .order_by(models.InventoryTransaction.transaction_date.asc())
        .all()
    )

    timeline = []
    balance = 0
    total_in = 0
    total_out = 0
    total_adj = 0
    first_in_date = None

    # CREATED event from audit log
    created_log = (
        db.query(models.AuditLog)
        .filter(models.AuditLog.action == "PRODUCT_CREATE", models.AuditLog.resource == urun.name_tr)
        .order_by(models.AuditLog.timestamp.asc())
        .first()
    )
    if created_log:
        timeline.append(LifecycleEvent(
            date=created_log.timestamp.isoformat(),
            stage="CREATED",
            actor=created_log.actor,
            quantity_delta=0,
            running_balance=0,
            source="Audit",
            notes=created_log.detail,
        ))

    for h in hareketler:
        if h.transaction_type == "IN":
            total_in += h.quantity
            balance += h.quantity
            stage = "RECEIVED"
            if first_in_date is None:
                first_in_date = h.transaction_date
        elif h.transaction_type == "OUT":
            total_out += h.quantity
            balance -= h.quantity
            stage = "CONSUMED"
        else:
            total_adj += h.quantity
            balance += h.quantity
            stage = "ADJUST"

        timeline.append(LifecycleEvent(
            date=h.transaction_date.isoformat() if h.transaction_date else "",
            stage=stage,
            actor=h.processed_by or "System",
            quantity_delta=h.quantity if h.transaction_type != "OUT" else -h.quantity,
            running_balance=balance,
            source=h.source or "Manuel",
            notes=h.notes,
        ))

    # Near-expiry / expired synthetic events
    now = datetime.datetime.utcnow().date()
    if urun.expiration_date:
        days_left = (urun.expiration_date - now).days
        if days_left <= 0:
            timeline.append(LifecycleEvent(
                date=urun.expiration_date.isoformat(), stage="EXPIRED", actor="System",
                quantity_delta=0, running_balance=balance, notes="SKT aşıldı / Expired",
            ))
        elif days_left <= SKT_ALERT_DAYS:
            timeline.append(LifecycleEvent(
                date=now.isoformat(), stage="NEAR_EXPIRY", actor="System",
                quantity_delta=0, running_balance=balance, notes=f"{days_left} gün kaldı / {days_left} days left",
            ))

    # Depleted event
    if urun.current_stock == 0 and total_in > 0:
        last_out = next((h for h in reversed(hareketler) if h.transaction_type == "OUT"), None)
        if last_out:
            timeline.append(LifecycleEvent(
                date=last_out.transaction_date.isoformat() if last_out.transaction_date else "",
                stage="DEPLETED", actor=last_out.processed_by or "System",
                quantity_delta=0, running_balance=0, notes="Stok tükendi / Stock depleted",
            ))

    # Summary metrics
    days_on_hand = None
    if first_in_date:
        days_on_hand = (datetime.datetime.utcnow() - first_in_date).days

    turnover_ratio = None
    if total_in > 0 and days_on_hand and days_on_hand > 0:
        avg_inventory = total_in / 2
        if avg_inventory > 0:
            turnover_ratio = round(total_out / avg_inventory, 2)

    remaining_shelf = None
    if urun.expiration_date:
        remaining_shelf = (urun.expiration_date - now).days

    summary = LifecycleSummary(
        sku=urun.sku,
        name_tr=urun.name_tr,
        name_en=urun.name_en,
        current_stock=urun.current_stock,
        total_received=total_in,
        total_consumed=total_out,
        total_adjusted=total_adj,
        days_on_hand=days_on_hand,
        turnover_ratio=turnover_ratio,
        remaining_shelf_life_days=remaining_shelf,
        expiration_date=urun.expiration_date.isoformat() if urun.expiration_date else None,
    )

    return ProductLifecycleResponse(summary=summary, timeline=timeline)


@app.get("/stok-yasam-dongusu", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def stok_yasam_dongusu(db: Session = Depends(get_db)):
    """Ürün bazında ilk giriş, son çıkış ve envanterde kalma süresi."""
    from sqlalchemy import func

    urunler = db.query(models.Product).all()
    sonuc = []
    for u in urunler:
        hareketler = (
            db.query(models.InventoryTransaction)
            .filter(models.InventoryTransaction.product_id == u.product_id)
            .order_by(models.InventoryTransaction.transaction_date.asc())
            .all()
        )
        ilk_giris = None
        son_cikis = None
        for h in hareketler:
            if h.transaction_type == "IN" and ilk_giris is None:
                ilk_giris = h.transaction_date
            if h.transaction_type == "OUT":
                son_cikis = h.transaction_date

        kalis_suresi = None
        kalis_gun = None
        if ilk_giris and son_cikis and son_cikis > ilk_giris:
            delta = son_cikis - ilk_giris
            kalis_gun = delta.days
            saat = delta.seconds // 3600
            dakika = (delta.seconds % 3600) // 60
            kalis_suresi = f"{kalis_gun}g {saat}s {dakika}dk"
        elif ilk_giris and not son_cikis:
            delta = datetime.datetime.utcnow() - ilk_giris
            kalis_gun = delta.days
            saat = delta.seconds // 3600
            dakika = (delta.seconds % 3600) // 60
            kalis_suresi = f"{kalis_gun}g {saat}s {dakika}dk"

        sonuc.append({
            "product_id": u.product_id,
            "sku": u.sku,
            "name_tr": u.name_tr,
            "name_en": u.name_en,
            "current_stock": u.current_stock,
            "ilk_giris": ilk_giris.isoformat() if ilk_giris else None,
            "son_cikis": son_cikis.isoformat() if son_cikis else None,
            "kalis_suresi": kalis_suresi,
            "kalis_gun": kalis_gun,
            "hala_depoda": son_cikis is None and ilk_giris is not None,
            "hareket_sayisi": len(hareketler),
        })
    return {"data": sonuc}


# --- GERÇEK ZAMANLI İŞLEM UÇ NOKTASI (ASYNC WEBSOCKET) ---
@app.post("/stok-hareketi")
async def stok_hareketi_kaydet(hareket: StockTransaction, background_tasks: BackgroundTasks, current_user: models.User = Depends(role_required(["Admin", "Barista", "Depo Müdürü"])), db: Session = Depends(get_db)):
    try:
        # 1. BİLGİSAYAR İŞ MANTIĞI: Barista yapıyorsa bekle, Yetkili yapıyorsa anında vur!
        islem_durumu = "BEKLEMEDE" if current_user.role == "Barista" else "ONAYLANDI"
        
        yeni_islem = models.InventoryTransaction(
            product_id=hareket.product_id,
            quantity=hareket.quantity,
            transaction_type=hareket.transaction_type,
            notes=hareket.notes,
            processed_by=current_user.username,
            status=islem_durumu
        )
        db.add(yeni_islem)
        
        # ONAYLI İSE ANINDA STOK GÜNCELLE
        if islem_durumu == "ONAYLANDI":
            urun = db.query(models.Product).filter(models.Product.product_id == hareket.product_id).first()
            if urun:
                if hareket.transaction_type.upper() == "IN": urun.current_stock += hareket.quantity
                else: urun.current_stock -= hareket.quantity
                
                # --- OTOMASYON TETİKLEYİCİLERİ ---
                if urun.current_stock <= urun.reorder_point:
                    # 1. Tedarikçiye E-posta
                    if urun.supplier and urun.supplier.contact_email:
                        background_tasks.add_task(send_supplier_email, urun.supplier.contact_email, urun.name_tr, urun.current_stock)
                    # 2. Webhook Gönderimi
                    background_tasks.add_task(trigger_stock_webhook, {
                        "event": "LOW_STOCK",
                        "product": urun.sku,
                        "current_stock": urun.current_stock,
                        "reorder_point": urun.reorder_point,
                        "timestamp": str(datetime.datetime.now())
                    })
        
        db.commit()
        
        # --- AUDIT LOG: STOCK MOVEMENT ---
        urun_adi = db.query(models.Product).filter(models.Product.product_id == hareket.product_id).first()
        _log_audit(
            db, actor=current_user.username, role=current_user.role,
            action=f"STOCK_{hareket.transaction_type.upper()}",
            resource=urun_adi.name_en if urun_adi else f"ID:{hareket.product_id}",
            detail=f"Qty: {hareket.quantity} | Status: {islem_durumu}",
        )
        
        # 2. CANLI YAYIN: Yöneticiye ping at!
        if islem_durumu == "BEKLEMEDE":
            await manager.broadcast(f"DiKKAT: {current_user.username} tarafından yeni bir onay talebi fırlatıldı!")
            return {"mesaj": "Talebiniz yönetici onayına sunuldu. Bildirim gönderildi."}
            
        return {"mesaj": "İşlem doğrudan onaylandı ve stok güncellendi."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/dashboard-ozet", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def dashboard_ozet(db: Session = Depends(get_db)):
    yatirim = db.query(func.sum(models.Product.current_stock * models.Product.unit_cost)).scalar() or 0
    kritik_sayisi = db.query(models.Product).filter(models.Product.current_stock <= models.Product.reorder_point).count()
    
    son_islemler = db.query(models.InventoryTransaction).order_by(models.InventoryTransaction.transaction_id.desc()).limit(5).all()
    sonuc_islemler = [{"islem": i.transaction_id, "tip": i.transaction_type, "adet": i.quantity, "urun_tr": i.product.name_tr, "urun_en": i.product.name_en} for i in son_islemler if i.product]

    return {
        "finansal_durum": {"toplam_yatirim_maliyeti": float(yatirim)},
        "kritik_stok_uyari_sayisi": kritik_sayisi,
        "son_islemler": sonuc_islemler
    }



@app.get("/tedarikci-siparis", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def tedarikci_siparis(db: Session = Depends(get_db)):
    kritik_urunler = db.query(models.Product).filter(models.Product.current_stock <= models.Product.reorder_point).all()
    siparisler = {}
    for u in kritik_urunler:
        ted = u.supplier.name if u.supplier else "Belirtilmemiş Tedarikçi"
        email = u.supplier.contact_email if u.supplier else "-"
        if ted not in siparisler:
            siparisler[ted] = {"email": email, "urunler": []}
        siparisler[ted]["urunler"].append({
            "urun_tr": u.name_tr, 
            "urun_en": u.name_en, 
            "mevcut_stok": u.current_stock, 
            "siparis_edilmesi_gereken": (u.reorder_point - u.current_stock) + 50
        })
    return {"bekleyen_siparis_listesi": siparisler}

@app.get("/skt-analizi", dependencies=[Depends(role_required(["Admin", "Depo Müdürü", "Barista"]))])
def skt_analizi(db: Session = Depends(get_db)):
    """Gerçek hayatta LIFO (Son Giren İlk Çıkar) kuralını uygulamak için 30 gün içinde bozulacak SKT riskli ürünleri saptar."""
    bugun = datetime.date.today()
    otuz_gun_sonra = bugun + datetime.timedelta(days=30)
    
    riskli_urunler = db.query(models.Product).filter(
        models.Product.expiration_date.isnot(None),
        models.Product.expiration_date <= otuz_gun_sonra,
        models.Product.current_stock > 0
    ).all()
    
    return {"skt_riskli_urunler": sorted([{
        "urun_tr": u.name_tr,
        "urun_en": u.name_en,
        "skt_tarihi": u.expiration_date.isoformat(),
        "kalan_gun": (u.expiration_date - bugun).days,
        "raf_konumu": u.warehouse_location,
        "stok": u.current_stock
    } for u in riskli_urunler], key=lambda x: x["kalan_gun"])}

@app.get("/fire-raporu", dependencies=[Depends(role_required(["Admin"]))])
def fire_raporu(db: Session = Depends(get_db)):
    fireler = db.query(models.InventoryTransaction).filter(
        models.InventoryTransaction.transaction_type == "OUT",
        or_(
            models.InventoryTransaction.notes.ilike('%fire%'),
            models.InventoryTransaction.notes.ilike('%bozuldu%'),
            models.InventoryTransaction.notes.ilike('%istisna%')
        )
    ).all()
    
    toplam = sum(float(f.quantity * (f.product.unit_cost if f.product else 0)) for f in fireler)
    return {"toplam_fire_zarari_tl": toplam, "fire_detaylari": [{"tarih": f.transaction_date, "not": f.notes} for f in fireler]}

@app.get("/talep-tahmini", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def talep_tahmini(db: Session = Depends(get_db)):
    otuz_gun_once = datetime.datetime.utcnow() - datetime.timedelta(days=REPORTING_WINDOW_DAYS)
    Gecmis_cikislar = db.query(
        models.InventoryTransaction.product_id,
        func.sum(models.InventoryTransaction.quantity).label("toplam")
    ).filter(
        models.InventoryTransaction.transaction_type == "OUT",
        models.InventoryTransaction.transaction_date >= otuz_gun_once
    ).group_by(models.InventoryTransaction.product_id).all()
    
    tahminler = []
    for (p_id, toplam_cikis) in Gecmis_cikislar:
        if not toplam_cikis: continue
        haftalik_tahmin = round((toplam_cikis / 30) * 7)
        urun = db.query(models.Product).filter(models.Product.product_id == p_id).first()
        tahminler.append({"urun_adi_tr": urun.name_tr if urun else f"ID:{p_id}", "urun_adi_en": urun.name_en if urun else f"ID:{p_id}", "gelecek_hafta_tahmini_talep": haftalik_tahmin})
        
    return {"haftalik_talep_tahmini": sorted(tahminler, key=lambda k: k["gelecek_hafta_tahmini_talep"], reverse=True)}

@app.get("/bekleyen-talepler", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def bekleyen_talepler(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    talepler = db.query(models.InventoryTransaction).filter(models.InventoryTransaction.status == "BEKLEMEDE").offset(skip).limit(limit).all()
    return {"talepler": [{
            "transaction_id": t.transaction_id,
            "transaction_date": t.transaction_date,
            "urun_adi_tr": t.product.name_tr if t.product else "?",
            "urun_adi_en": t.product.name_en if t.product else "?",
            "transaction_type": t.transaction_type,
            "quantity": t.quantity,
            "talep_eden": t.processed_by
        } for t in talepler]}

@app.put("/talep-yanitla/{islem_id}")
async def talep_yanitla(islem_id: int, yanit: TalepYaniti, background_tasks: BackgroundTasks, current_user: models.User = Depends(role_required(["Admin", "Depo Müdürü"])), db: Session = Depends(get_db)):
    talep = db.query(models.InventoryTransaction).filter(models.InventoryTransaction.transaction_id == islem_id).first()
    if not talep:
        raise HTTPException(status_code=404, detail="Talep bulunamadı.")
    if talep.status != "BEKLEMEDE":
        raise HTTPException(status_code=409, detail="Bu talep zaten yanıtlanmış.")
    
    talep.status = yanit.yeni_durum.upper()
    talep.notes = f"{(talep.notes or '')} | Yanıtlayan: {current_user.username}"
    
    if talep.status == "ONAYLANDI":
        urun = talep.product
        if urun:
            if talep.transaction_type.upper() == "IN": urun.current_stock += talep.quantity
            else: urun.current_stock -= talep.quantity
            
            # --- OTOMASYON TETİKLEYİCİLERİ ---
            if urun.current_stock <= urun.reorder_point:
                if urun.supplier and urun.supplier.contact_email:
                    background_tasks.add_task(send_supplier_email, urun.supplier.contact_email, urun.name_tr, urun.current_stock)
                background_tasks.add_task(trigger_stock_webhook, {
                    "event": "LOW_STOCK_AFTER_APPROVAL",
                    "product": urun.sku,
                    "current_stock": urun.current_stock,
                    "reorder_point": urun.reorder_point,
                    "timestamp": str(datetime.datetime.now())
                })
                
    db.commit()
    _log_audit(
        db, actor=current_user.username, role=current_user.role,
        action=f"APPROVE_{yanit.yeni_durum.upper()}",
        resource=talep.product.name_en if talep.product else f"ID:{islem_id}",
        detail=f"Tx #{islem_id}",
    )
    return {"mesaj": f"Talep {talep.status} olarak işlendi."}




@app.post("/tedarik-siparis-onayla", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
async def tedarik_siparis_onayla(req: SupplyApprovalRequest, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """Yönetici onayıyla tedarik siparişi oluşturur. Otomatik sipariş YOKTUR."""
    try:
        urun = db.query(models.Product).filter(models.Product.product_id == req.product_id).first()
        if not urun:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı.")
        onay = models.SupplyOrderApproval(
            product_id=req.product_id,
            quantity=req.quantity,
            approved_by=current_user.username,
            supplier_name=req.supplier_name or (urun.supplier.name if urun.supplier else "Belirtilmemiş"),
            status="ORDERED"
        )
        db.add(onay)
        db.add(models.AuditLog(
            actor=current_user.username, role=current_user.role,
            action="SUPPLY_ORDER_APPROVED",
            resource=urun.name_tr,
            detail=f"{req.quantity} adet sipariş onaylandı.",
        ))
        db.commit()
        return {"mesaj": f"{urun.name_tr} için {req.quantity} adet tedarik siparişi onaylandı.", "onay_id": onay.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))




@app.get("/sevk-raporu", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def sevk_raporu(db: Session = Depends(get_db)):
    # N+1 sorgu yerine tek JOIN ile top-5 çıkış ürünü
    en_cok = db.query(
        models.Product.name_tr,
        models.Product.name_en,
        func.sum(models.InventoryTransaction.quantity).label("toplam")
    ).join(
        models.InventoryTransaction,
        models.Product.product_id == models.InventoryTransaction.product_id
    ).filter(
        models.InventoryTransaction.transaction_type == "OUT"
    ).group_by(
        models.Product.product_id, models.Product.name_tr, models.Product.name_en
    ).order_by(
        func.sum(models.InventoryTransaction.quantity).desc()
    ).limit(5).all()

    pasta = [{"name_tr": row.name_tr, "name_en": row.name_en, "toplam_sevk": float(row.toplam or 0)} for row in en_cok]

    yedi_gecmis = datetime.datetime.utcnow() - datetime.timedelta(days=7)
    trend_data = db.query(
        func.date(models.InventoryTransaction.transaction_date).label("dt"),
        func.sum(models.InventoryTransaction.quantity).label("toplam")
    ).filter(
        models.InventoryTransaction.transaction_type == "OUT",
        models.InventoryTransaction.transaction_date >= yedi_gecmis
    ).group_by(func.date(models.InventoryTransaction.transaction_date)).all()

    return {"grafik_1_pasta": {"veriler": pasta}, "grafik_2_cizgi": {"veriler": [{"tarih": str(t.dt), "gunluk_cikis_adeti": float(t.toplam or 0)} for t in trend_data]}}


# ==========================================
# --- AUDIT TRAIL (GÜVENLİK DENETİM KAYITLARI) ---
# ==========================================
@app.get("/audit-logs", dependencies=[Depends(role_required(["Admin"]))])
def audit_logs(skip: int = 0, limit: int = 50, db: Session = Depends(get_db)):
    logs = db.query(models.AuditLog)\
             .filter(models.AuditLog.is_archived == 0)\
             .order_by(models.AuditLog.timestamp.desc())\
             .offset(skip).limit(limit).all()
    total = db.query(func.count(models.AuditLog.id)).filter(models.AuditLog.is_archived == 0).scalar()
    return {
        "total": total,
        "logs": [{
            "id": l.id,
            "actor": l.actor,
            "role": l.role,
            "action": l.action,
            "resource": l.resource or "-",
            "detail": l.detail or "-",
            "ip_address": l.ip_address or "-",
            "timestamp": l.timestamp.isoformat() if l.timestamp else "-"
        } for l in logs]
    }


# ==========================================
# --- 2. ANALİTİK & RAPORLAMA MODÜLÜ ---
# ==========================================

@app.get("/rapor/kar-marji", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def kar_marji_analizi(db: Session = Depends(get_db)):
    """
    Her SKU için kâr marjı, EWMA (ağırlıklı hareketli ortalama) tüketim tahmini
    ve stok yeterliliği analizi. Sade SQL + Python ile hesaplanır.
    """
    urunler = db.query(models.Product).filter(models.Product.current_stock > 0).all()
    otuz_gun_once = datetime.datetime.utcnow() - datetime.timedelta(days=REPORTING_WINDOW_DAYS)

    # Son 30 günlük çıkış hacimleri — tek sorguda
    cikis_map = {
        row.product_id: float(row.toplam or 0)
        for row in db.query(
            models.InventoryTransaction.product_id,
            func.sum(models.InventoryTransaction.quantity).label("toplam")
        ).filter(
            models.InventoryTransaction.transaction_type == "OUT",
            models.InventoryTransaction.transaction_date >= otuz_gun_once
        ).group_by(models.InventoryTransaction.product_id).all()
    }

    sonuclar = []
    for u in urunler:
        marj_tl   = round(u.unit_price - u.unit_cost, 2)
        marj_yuzde = round((marj_tl / u.unit_price * 100) if u.unit_price > 0 else 0, 1)
        toplam_kar = round(marj_tl * u.current_stock, 2)

        # EWMA (α=EWMA_ALPHA): son 30 günlük günlük ortalama tüketim
        aylik_cikis = cikis_map.get(u.product_id, 0.0)
        gunluk_ort  = aylik_cikis / REPORTING_WINDOW_DAYS
        ewma_talep  = round(gunluk_ort * EWMA_ALPHA + gunluk_ort * (1 - EWMA_ALPHA), 2)
        kac_gun_yeter = round(u.current_stock / ewma_talep, 1) if ewma_talep > 0 else None

        sonuclar.append({
            "product_id":     u.product_id,
            "sku":            u.sku,
            "name_tr":        u.name_tr,
            "name_en":        u.name_en,
            "unit_cost":      u.unit_cost,
            "unit_price":     u.unit_price,
            "marj_tl":        marj_tl,
            "marj_yuzde":     marj_yuzde,
            "toplam_potansiyel_kar_tl": toplam_kar,
            "current_stock":  u.current_stock,
            "ewma_gunluk_talep": ewma_talep,
            "stok_kac_gun_yeter": kac_gun_yeter,
            "kritik": kac_gun_yeter is not None and kac_gun_yeter < 7
        })

    # Kâr potansiyeline göre sırala
    sonuclar.sort(key=lambda x: x["toplam_potansiyel_kar_tl"], reverse=True)
    toplam_potansiyel = round(sum(s["toplam_potansiyel_kar_tl"] for s in sonuclar), 2)

    logger.info(f"Kar Marjı Analizi: {len(sonuclar)} SKU işlendi.")
    return {
        "toplam_potansiyel_kar_tl": toplam_potansiyel,
        "urun_sayisi": len(sonuclar),
        "analiz": sonuclar
    }


@app.get("/rapor/excel", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def excel_rapor_indir(db: Session = Depends(get_db)):
    """Tüm ürün envanterini + kâr marjı hesaplamalarını .xlsx olarak indirir."""
    wb = openpyxl.Workbook()

    # ── SAYFA 1: ÜRÜN ENVANTERİ ──
    ws1 = wb.active
    ws1.title = "Envanter"

    baslik_font  = Font(bold=True, color="FFFFFF", size=11)
    baslik_dolu  = PatternFill("solid", fgColor="1E293B")
    orta_hizala  = Alignment(horizontal="center")

    sutunlar = ["SKU", "Ürün (TR)", "Ürün (EN)", "Kategori", "Mevcut Stok",
                "Birim Maliyet (₺)", "Birim Fiyat (₺)", "Kâr Marjı (₺)",
                "Kâr Marjı (%)", "Toplam Potansiyel Kâr (₺)", "SKT", "Raf"]
    ws1.append(sutunlar)
    for col_idx, _ in enumerate(sutunlar, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font   = baslik_font
        cell.fill   = baslik_dolu
        cell.alignment = orta_hizala

    urunler = db.query(models.Product).all()
    for u in urunler:
        marj_tl    = round(u.unit_price - u.unit_cost, 2)
        marj_yuzde = round((marj_tl / u.unit_price * 100) if u.unit_price > 0 else 0, 1)
        ws1.append([
            u.sku,
            u.name_tr,
            u.name_en,
            u.category.name_tr if u.category else "-",
            u.current_stock,
            u.unit_cost,
            u.unit_price,
            marj_tl,
            marj_yuzde,
            round(marj_tl * u.current_stock, 2),
            u.expiration_date.isoformat() if u.expiration_date else "-",
            u.warehouse_location or "-"
        ])

    # Sütun genişliklerini otomatik ayarla
    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── SAYFA 2: SON 30 GÜN STK HAREKETLERİ ──
    ws2 = wb.create_sheet("Stok Hareketleri")
    ws2.append(["Tarih", "Ürün", "Tip", "Miktar", "Durum", "İşleyen", "Not"])
    ws2.cell(row=1, column=1).font = baslik_font

    otuz_gun_once = datetime.datetime.utcnow() - datetime.timedelta(days=REPORTING_WINDOW_DAYS)
    hareketler = db.query(models.InventoryTransaction)\
        .filter(models.InventoryTransaction.transaction_date >= otuz_gun_once)\
        .order_by(models.InventoryTransaction.transaction_date.desc()).all()
    for h in hareketler:
        ws2.append([
            h.transaction_date.strftime("%Y-%m-%d %H:%M") if h.transaction_date else "-",
            h.product.name_tr if h.product else f"ID:{h.product_id}",
            h.transaction_type,
            h.quantity,
            h.status,
            h.processed_by,
            h.notes or "-"
        ])

    # ── SAYFA 3: AUDIT LOG (Son 100) ──
    ws3 = wb.create_sheet("Denetim Kaydı")
    ws3.append(["Zaman", "Kullanıcı", "Rol", "İşlem", "Kaynak", "Detay", "IP"])
    loglar = db.query(models.AuditLog).order_by(models.AuditLog.timestamp.desc()).limit(100).all()
    for l in loglar:
        ws3.append([
            l.timestamp.strftime("%Y-%m-%d %H:%M:%S") if l.timestamp else "-",
            l.actor, l.role, l.action,
            l.resource or "-", l.detail or "-", l.ip_address or "-"
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    tarih = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"envanter_raporu_{tarih}.xlsx"

    logger.info(f"Excel raporu oluşturuldu: {filename}")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/rapor/pdf", dependencies=[Depends(role_required(["Admin", "Depo Müdürü"]))])
def pdf_rapor_indir(db: Session = Depends(get_db)):
    """Yönetici özet raporunu (Dashboard KPI + Kritik Stoklar + Kâr Marjı Top-5) PDF olarak indirir."""
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    KOYU   = colors.HexColor("#0f172a")
    YESIL  = colors.HexColor("#34d399")
    KIRMIZI = colors.HexColor("#fb7185")
    GRIS   = colors.HexColor("#94a3b8")

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  textColor=KOYU, fontSize=18, spaceAfter=4)
    alt_style   = ParagraphStyle("alt", parent=styles["Normal"],
                                  textColor=GRIS, fontSize=9, spaceAfter=12)
    h2_style    = ParagraphStyle("h2", parent=styles["Heading2"],
                                  textColor=KOYU, fontSize=13, spaceBefore=14, spaceAfter=6)

    # Başlık
    story.append(Paragraph("🏛️ Akıllı Kafe Envanter Sistemi", title_style))
    story.append(Paragraph(f"Yönetici Raporu — {datetime.datetime.now().strftime('%d %B %Y, %H:%M')}", alt_style))
    story.append(HRFlowable(width="100%", thickness=1, color=YESIL, spaceAfter=14))

    # KPI Kutuları
    yatirim    = db.query(func.sum(models.Product.current_stock * models.Product.unit_cost)).scalar() or 0
    kritik_say = db.query(models.Product).filter(models.Product.current_stock <= models.Product.reorder_point).count()
    urun_say   = db.query(models.Product).count()

    story.append(Paragraph("📊 Finansal Özet", h2_style))
    kpi_data = [
        ["Metrik", "Değer"],
        ["Toplam Envanter Yatırım Maliyeti", f"₺{float(yatirim):,.2f}"],
        ["Toplam Aktif SKU Sayısı",          str(urun_say)],
        ["Kritik Stok Uyarısı (≤ Reorder Point)", str(kritik_say)],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[10*cm, 5*cm])
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), KOYU),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#f8fafc"), colors.white]),
        ("GRID",        (0,0), (-1,-1), 0.5, GRIS),
        ("ALIGN",       (1,0), (1,-1), "RIGHT"),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("RIGHTPADDING",(0,0), (-1,-1), 8),
        ("TOPPADDING",  (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.5*cm))

    # Top-5 Kâr Marjı
    story.append(Paragraph("💰 En Yüksek Kâr Potansiyeli — Top 5 SKU", h2_style))
    urunler = db.query(models.Product).filter(models.Product.unit_price > 0).all()
    top5 = sorted(urunler,
                  key=lambda u: (u.unit_price - u.unit_cost) * u.current_stock,
                  reverse=True)[:5]
    kar_data = [["SKU", "Ürün", "Marj (%)", "Stok", "Potansiyel Kâr (₺)"]]
    for u in top5:
        marj_tl = u.unit_price - u.unit_cost
        marj_yuzde = round((marj_tl / u.unit_price * 100) if u.unit_price > 0 else 0, 1)
        kar_data.append([
            u.sku, u.name_tr[:30],
            f"%{marj_yuzde}",
            str(u.current_stock),
            f"₺{round(marj_tl * u.current_stock, 2):,.2f}"
        ])
    kar_tbl = Table(kar_data, colWidths=[2.5*cm, 6*cm, 2*cm, 1.5*cm, 3.5*cm])
    kar_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), YESIL),
        ("TEXTCOLOR",   (0,0), (-1,0), KOYU),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#f0fdf4"), colors.white]),
        ("GRID",        (0,0), (-1,-1), 0.5, GRIS),
        ("ALIGN",       (2,0), (-1,-1), "CENTER"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))
    story.append(kar_tbl)
    story.append(Spacer(1, 0.5*cm))

    # Kritik Stok Uyarıları
    kritik_urunler = db.query(models.Product)\
        .filter(models.Product.current_stock <= models.Product.reorder_point).limit(10).all()
    if kritik_urunler:
        story.append(Paragraph("⚠️ Kritik Stok Uyarıları", h2_style))
        uyari_data = [["SKU", "Ürün", "Mevcut", "Eşik"]]
        for u in kritik_urunler:
            uyari_data.append([u.sku, u.name_tr[:35], str(u.current_stock), str(u.reorder_point)])
        uyari_tbl = Table(uyari_data, colWidths=[2.5*cm, 8*cm, 2.5*cm, 2.5*cm])
        uyari_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), KIRMIZI),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#fff1f2"), colors.white]),
            ("GRID",        (0,0), (-1,-1), 0.5, GRIS),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ]))
        story.append(uyari_tbl)

    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRIS))
    story.append(Paragraph(
        f"<font color='#94a3b8' size='8'>Bu rapor otomatik olarak üretilmiştir. "
        f"Akıllı Kafe Envanter Sistemi V2 — {datetime.datetime.now().strftime('%Y')}</font>",
        styles["Normal"]
    ))

    doc.build(story)
    buffer.seek(0)
    tarih = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"yonetici_raporu_{tarih}.pdf"

    logger.info(f"PDF raporu oluşturuldu: {filename}")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

